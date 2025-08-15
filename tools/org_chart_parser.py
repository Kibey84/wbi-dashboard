import os
import re
import json
import asyncio
import logging
from typing import Optional, List, Dict, Any

import pandas as pd
import fitz  # PyMuPDF

from openai import AsyncAzureOpenAI
from openai.types.chat import ChatCompletionMessageParam
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ORGCHART_CONCURRENCY = int(os.getenv("ORGCHART_CONCURRENCY", "4"))

# ------------------------------------------------------------------------------
# Logging helper
# ------------------------------------------------------------------------------
logger = logging.getLogger("wbi") if logging.getLogger().hasHandlers() else None

def _log(level: str, msg: str, *args):
    if logger:
        getattr(logger, level)(msg, *args)
    else:
        print(f"[{level.upper()}] " + (msg % args if args else msg))

# ------------------------------------------------------------------------------
# JSON helpers
# ------------------------------------------------------------------------------
_JSON_SNIPPET = re.compile(r"(\[.*?\]|\{.*?\})", re.DOTALL)

def _load_json_lenient(text: str) -> Optional[Any]:
    """Try to load JSON; if that fails, extract the first {…} or […] block and retry."""
    if not text:
        return None
    try:
        return json.loads(text)
    except Exception:
        match = _JSON_SNIPPET.search(text)
        if match:
            try:
                return json.loads(match.group(1))
            except Exception:
                return None
        return None

# ------------------------------------------------------------------------------
# Azure OpenAI call with JSON-mode fallback + light retries
# ------------------------------------------------------------------------------
async def call_azure_ai(
    client: AsyncAzureOpenAI,
    prompt_text: str,
    *,
    system_prompt: str = "You are an expert data entry assistant for org charts.",
    max_tokens: int = 4096,
    temperature: float = 0.1,
    retries: int = 2,
    retry_base_delay: float = 0.8,
) -> Optional[str]:
    """
    Call AOAI:
      1) Try JSON mode
      2) On failure, retry without response_format
      3) Light retries/backoff for transient issues
    Returns raw text (may be JSON or prose containing JSON).
    """
    deployment_name = (os.getenv("AZURE_OPENAI_DEPLOYMENT") or "").strip()
    if not deployment_name:
        _log("error", "AZURE_OPENAI_DEPLOYMENT missing")
        return None

    messages: List[ChatCompletionMessageParam] = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": prompt_text},
    ]

    try_json_mode_first = True
    last_err: Optional[Exception] = None

    for attempt in range(retries + 1):
        try:
            if try_json_mode_first:
                try:
                    resp = await client.chat.completions.create(
                        model=deployment_name,
                        messages=messages,
                        max_tokens=max_tokens,
                        temperature=temperature,
                        response_format={"type": "json_object"},
                    )
                except Exception as e:
                    _log("warning", "JSON mode failed; retrying without it: %s", str(e))
                    try_json_mode_first = False
                    resp = await client.chat.completions.create(
                        model=deployment_name,
                        messages=messages,
                        max_tokens=max_tokens,
                        temperature=temperature,
                    )
            else:
                resp = await client.chat.completions.create(
                    model=deployment_name,
                    messages=messages,
                    max_tokens=max_tokens,
                    temperature=temperature,
                )

            content = (resp.choices[0].message.content or "").strip() if resp.choices else ""
            return content or ""
        except Exception as e:
            last_err = e
            if attempt < retries:
                delay = retry_base_delay * (2 ** attempt)
                _log("warning", "AOAI call failed (attempt %d/%d): %s; retrying in %.1fs",
                     attempt + 1, retries + 1, str(e), delay)
                await asyncio.sleep(delay)

    _log("error", "AOAI call failed after retries: %s", str(last_err) if last_err else "unknown error")
    return None

# ------------------------------------------------------------------------------
# Chunk parsing
# ------------------------------------------------------------------------------
async def parse_text_chunk(client: AsyncAzureOpenAI, text_chunk: str) -> List[dict]:
    """
    Ask the model to extract units into a JSON array of:
      [{ "name": "", "leader": "", "title": "", "location": "" }, ...]
    Returns [] on any problem; never raises.
    """
    prompt = f"""
From the text below, extract a list of all distinct organizational units.

Rules:
- Return a JSON array where each object has keys: "name", "leader", "title", "location".
- Use "" for any missing field.
- Do NOT invent or infer information not present in the text.
- Output only the JSON array.

---
TEXT TO PARSE:
{text_chunk}
"""
    raw = await call_azure_ai(client, prompt)
    if not raw:
        _log("warning", "No response received for a text chunk.")
        return []

    data = _load_json_lenient(raw)
    if isinstance(data, list):
        out: List[dict] = []
        for item in data:
            if not isinstance(item, dict):
                continue
            out.append({
                "name": item.get("name", "") or "",
                "leader": item.get("leader", "") or "",
                "title": item.get("title", "") or "",
                "location": item.get("location", "") or "",
            })
        return out

    _log("warning", "Model did not return a JSON array. Sample: %s", raw[:300])
    return []

# ------------------------------------------------------------------------------
# Excel output
# ------------------------------------------------------------------------------
def save_and_format_excel(df: pd.DataFrame, output_directory: str, output_filename: str) -> None:
    """Saves the DataFrame to a formatted Excel file, grouped by location."""
    if df.empty:
        _log("warning", "DataFrame empty; skipping Excel save.")
        return

    os.makedirs(output_directory, exist_ok=True)
    full_path = os.path.join(output_directory, output_filename)

    df["Location"] = df["Location"].fillna("Unknown Location")
    df_sorted = df.sort_values(by="Location")

    wb = Workbook()
    ws = wb.active
    assert isinstance(ws, Worksheet)
    ws.title = "OrgChart"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    location_font = Font(bold=True, color="FFFFFF")
    location_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    row_fill_1 = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    headers = ["Unit", "Leader", "Title", "Location"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_location = None
    row_color_index = 0
    for _, row in df_sorted.iterrows():
        if row["Location"] != current_location:
            current_location = row["Location"]
            ws.append([current_location] + [""] * (len(headers) - 1))
            location_header_row = ws.max_row
            ws.merge_cells(
                start_row=location_header_row, start_column=1,
                end_row=location_header_row, end_column=len(headers)
            )
            cell = ws.cell(row=location_header_row, column=1)
            cell.font = location_font
            cell.fill = location_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            row_color_index = 0

        data_row = [row.get(col, "") for col in ["Unit", "Leader", "Title", "Location"]]
        ws.append(data_row)

        current_fill = row_fill_1 if (row_color_index % 2 == 0) else row_fill_2
        for c in ws[ws.max_row]:
            c.fill = current_fill
        row_color_index += 1

    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            v = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    wb.save(full_path)
    _log("info", "Saved formatted Excel: %s", full_path)

# ------------------------------------------------------------------------------
# PDF processing
# ------------------------------------------------------------------------------
async def _async_process_pdf(
    pdf_bytes: bytes,
    concurrency: int = 4,
) -> List[dict]:
    """Core async logic for processing the PDF with capped concurrency."""
    all_units: List[dict] = []

    endpoint = (os.getenv("AZURE_OPENAI_ENDPOINT") or "").strip()
    key = (os.getenv("AZURE_OPENAI_KEY") or "").strip()
    if not endpoint or not key:
        _log("error", "Azure OpenAI credentials not found (AZURE_OPENAI_ENDPOINT/KEY).")
        return []

    client = AsyncAzureOpenAI(
        azure_endpoint=endpoint,
        api_key=key,
        api_version="2024-02-01",
        timeout=90.0,
    )

    sem = asyncio.Semaphore(ORGCHART_CONCURRENCY)
    async def guarded_parse(text):
        async with sem:
            return await parse_text_chunk(client, text)

    async def _guarded(text: str) -> List[dict]:
        async with sem:
            return await parse_text_chunk(client, text)

    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            page_texts: List[str] = []
            for i, page in enumerate(doc):  # type: ignore
                text = page.get_text("text")
                if text.strip():
                    page_texts.append(text)
                else:
                    _log("info", "Page %d contains no extractable text.", i + 1)

            _log("info", "Processing %d pages with concurrency=%d", len(page_texts), concurrency)
            tasks = [asyncio.create_task(_guarded(t)) for t in page_texts]
            for coro in asyncio.as_completed(tasks):
                try:
                    result_list = await coro
                    all_units.extend(result_list)
                except Exception as e:
                    _log("warning", "Chunk task failed: %s", str(e))
    finally:
        await client.close()

    return all_units

def process_uploaded_pdf(uploaded_file, output_directory: str) -> Optional[str]:
    """Synchronous wrapper that runs the entire async PDF processing workflow."""
    try:
        if not hasattr(uploaded_file, "filename") or not uploaded_file.filename:
            _log("error", "Uploaded file has no filename.")
            return None

        pdf_bytes = uploaded_file.read()
        if not pdf_bytes:
            _log("error", "Uploaded file is empty.")
            return None

        # Run the async pipeline
        all_units = asyncio.run(
            _async_process_pdf(
                pdf_bytes,
                concurrency=int(os.getenv("ORGCHART_CONCURRENCY", "4")),
            )
        )

        if not all_units:
            _log("warning", "AI returned no data from any page.")
            return None

        df = pd.DataFrame(all_units).rename(columns={
            "name": "Unit",
            "leader": "Leader",
            "title": "Title",
            "location": "Location",
        })

        for col in ["Unit", "Leader", "Title", "Location"]:
            if col not in df.columns:
                df[col] = ""

        original_filename = os.path.splitext(uploaded_file.filename)[0]
        output_filename = f"AI_Formatted_{original_filename}.xlsx"
        save_and_format_excel(df, output_directory, output_filename)
        return output_filename

    except Exception as e:
        _log("error", "PDF Processing Error: %s", str(e))
        return None
